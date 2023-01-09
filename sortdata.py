from openpyxl import load_workbook

file = "test.xlsx" #load the work book
wb_obj = load_workbook(filename = 'all_payslips.xlsx')
wsheet = wb_obj['all_payslips']
dataDict = {}

data={}

#ID	CompanyCode	FISCALYEAR	Mon	MonthName	CODENO	Group	GroupNo	Position	Desc	Value	Name	Department	Description	BANKNAME	BRANCH	PayMethod	NPFNUMBER	JobTitle	msg1	msg2	msg3	Wstation	WStation_Name	Hired_Date	STATUS	upsize_ts
#0	 1	              2   	3	       4	5	      6	       7    	8	     9	        10	11	      12	     13	           14	      15	16	           17	       18	     19	      20	21	      22	      23	            24	       25	26


#id	empID	salary	allowances	pension_employee	pension_employer	medical_employee	medical_employer	taxdue	meals	department	position	branch	pension_fund	membership_no	sdl	wcf	less_takehome 0-Complete Take Home	due_date	payroll_date	bank	bank_branch	account_no	created_at	updated_at
#id	empID	salary	allowances	pension_employee	pension_employer	medical_employee	medical_employer	taxdue	meals	department	position	branch	pension_fund	membership_no	sdl	wcf	less_takehome	due_date	payroll_date	bank	bank_branch	account_no	created_at	updated_at
#1   	2	 3	     4	         5	                  6	                  7	                    8	              9	     10	     11	         12	          13	    14	              15         16	 17 	18	        19	        20	            21	        22	                    23	        24	25
k=0
r=0
temp=[]
while r<30:
    temp.append('')
    r+=1
    # print(r)
for key, *values in wsheet.iter_rows():
    temp={}
    if(k>=1):
        # date='25/'+str(values[2].value)+'/'+str(values[1].value)
        date='25-'+str(values[2].value)+'-'+str(values[1].value)
        # print(date)
        # for value in values:
        if str(values[4].value)+date not in data:

            data[str(values[4].value)+date]=temp
            # print("Here")
            

        # dataDict[key.value] = [v.value for v in values]
        # print(dataDict[key.value])


            data[str(values[4].value)+date]['taxable_gross']='0'  #taxable gross

            data[str(values[4].value)+date]['gross_pay']='0'  #gross pay

            data[str(values[4].value)+date]['total_deduction']='0'  #Total Deduction

            data[str(values[4].value)+date]['tax_free_pension']='0'  #Total Deduction

            data[str(values[4].value)+date]['take_home']='0'  #less_takehome

            data[str(values[4].value)+date]['take_home']='0'  #less_takehome

            data[str(values[4].value)+date]['salary']='0'  #salary

            data[str(values[4].value)+date]['net_tax']='0'  #Net Tax

            data[str(values[4].value)+date]['paye']='0'  #salary

            data[str(values[4].value)+date]['salary']='0'  #salary

            data[str(values[4].value)+date]['gross_salary']='0'  #salary

            data[str(values[4].value)+date]['salary']='0'  #salary

            data[str(values[4].value)+date]['o/t']='0'  #o/t

            data[str(values[4].value)+date]['arrears']='0'  #arrears

            data[str(values[4].value)+date]['leave_balance']='0'  #leave balance

            data[str(values[4].value)+date]['teller_back_office']='0'  #leave balance

            data[str(values[4].value)+date]['helsb']='0'  #helsb

            data[str(values[4].value)+date]['helsb']='0'  #helsb

            data[str(values[4].value)+date]['helsb']='0'  #helsb

            data[str(values[4].value)+date]['helsb']='0'  #helsb

            data[str(values[4].value)+date]['helsb']='0'  #helsb

            data[str(values[4].value)+date]['helsb']='0'  #helsb

            data[str(values[4].value)+date]['helsb']='0'  #helsb

            data[str(values[4].value)+date]['task_allowances']='0'  #task_allowances

            data[str(values[4].value)+date]['helsb']='0'  #helsb

            data[str(values[4].value)+date]['helsb']='0'  #helsb

            data[str(values[4].value)+date]['helsb']='0'  #helsb

            data[str(values[4].value)+date]['helsb']='0'  #helsb

            data[str(values[4].value)+date]['normal_overtime']='0'  #normal_overtime

            data[str(values[4].value)+date]['sunday_overtime']='0'  #sunday_overtime

            data[str(values[4].value)+date]['weekday_overtime']='0'  #weekday_overtime

            data[str(values[4].value)+date]['sunday_overtime2']='0'  #sunday_overtime2

            data[str(values[4].value)+date]['rounding']='0'  #rounding

            data[str(values[4].value)+date]['helsb']='0'  #helsb

            data[str(values[4].value)+date]['acting_allowance']='0'  #acting_allowance

            data[str(values[4].value)+date]['helsb']='0'  #helsb

            data[str(values[4].value)+date]['disturbance']='0'  #helsb

            data[str(values[4].value)+date]['shift_allowance']='0'  #helsb

            data[str(values[4].value)+date]['transport_allow']='0'  #transport_allow

            data[str(values[4].value)+date]['helsb']='0'  #helsb

            data[str(values[4].value)+date]['unpdaid_leave']='0'  #unpdaid_leave

            data[str(values[4].value)+date]['o/stand_leave']='0'  #o/stand_leave

            data[str(values[4].value)+date]['leave_allowances']='0'  #allowances

            data[str(values[4].value)+date]['helsb']='0'  #helsb

            data[str(values[4].value)+date]['house_rent']='0'  #house rent

            data[str(values[4].value)+date]['air_ticket']='0'  #air_ticket

            data[str(values[4].value)+date]['hse_loan']='0'  #hse_loan

            data[str(values[4].value)+date]['vehicle_benefit']='0'  #vehicle_benefit

            data[str(values[4].value)+date]['bonus']='0'  #bonus

            data[str(values[4].value)+date]['advance']='0'  #advance

            data[str(values[4].value)+date]['mid_month_advance']='0'  #mid_month_advance

            data[str(values[4].value)+date]['sunday_overtime']='0'  #sunday_overtime

            data[str(values[4].value)+date]['service_award']='0'  #service_award

            data[str(values[4].value)+date]['share_all_payment']='0'  #share_all_payment

            data[str(values[4].value)+date]['salary']='0'  #salary

            data[str(values[4].value)+date]['loan_benefit']='0'  #loan_benefit

            data[str(values[4].value)+date]['level_i/o']='0'  #loan_benefit

            data[str(values[4].value)+date]['long_service']='0'  #loan_benefit

            data[str(values[4].value)+date]['recognition']='0'  #loan_benefit

            data[str(values[4].value)+date]['tra_cost']='0'  #loan_benefit

            data[str(values[4].value)+date]['unpaid_days']='0'  #loan_benefit

            data[str(values[4].value)+date]['employer_contribution']=0
            
            data[str(values[4].value)+date]['employee_contribution']=0
            data[str(values[4].value)+date]['membership_no']=0

        # data[str(values[4].value)+date]['pension_employee']=values[2].value  #pension_employee

        # data[str(values[4].value)+date]['pension_employer']=values[2].value  #pension_employer

        # data[str(values[4].value)+date]['medical_employee']=values[2].value  #medical_employee

        # data[str(values[4].value)+date]['medical_employer']=values[2].value  #medical_employer

        # data[str(values[4].value)+date]['taxdue']=values[2].value  #taxdue

        # data[str(values[4].value)+date]['meals']=values[2].value  #meals

        # data[str(values[4].value)+date]['department']=values[2].value  #department

        # data[str(values[4].value)+date]['position']=values[2].value  #position

        # data[str(values[4].value)+date]['branch']=values[2].value  #branch

        # data[str(values[4].value)+date]['pension_fund']=values[2].value  #pension_fund

        # data[str(values[4].value)+date]['membership_no']=values[2].value  #membership_no

        # data[str(values[4].value)+date]['sdl']=values[2].value  #sdl

        # data[str(values[4].value)+date]['wcf']=values[2].value  #wcf

        # data[str(values[4].value)+date]['less_takehome']=values[2].value  #less_takehome

        data[str(values[4].value)+date]['date']=date  #due_date

        # data[str(values[4].value)+date]['payroll_date']=date  #payroll_date

    #     data[str(values[4].value)+date]['bank']=values[2].value  #bank

    #     data[str(values[4].value)+date]['bank_branch']=values[2].value  #bank_branch

    #     data[str(values[4].value)+date]['account_no']=values[2].value  #account_no

    #     data[str(values[4].value)+date]['created_at']=values[2].value  #created_at


        # for key in dataDict:
        #     print(dataDict[key][3])
        data[str(values[4].value)+date]['empID']=values[4].value  #empID
        # print(values[8].value)

        if(values[8].value== "    Desc".replace(" ", "")):
            j=0
        # elif(str(values[8].value).replace(" ", "") == "9".replace(" ", "")):

        

        elif(str(values[8].value).replace(" ", "") == "Taxable Gross".replace(" ", "")):
             data[str(values[4].value)+date]['taxable_gross']=values[9].value  #taxable gross

        elif(str(values[8].value).replace(" ", "") == "Gross pay".replace(" ", "")):
            data[str(values[4].value)+date]['gross_pay']=values[9].value  #gross pay

        elif(str(values[8].value).replace(" ", "") == "Total Deduction".replace(" ", "")):
             data[str(values[4].value)+date]['total_deduction']=values[9].value  #Total Deduction

        elif(str(values[8].value).replace(" ", "") == "Less: Tax free Pension".replace(" ", "")):
            data[str(values[4].value)+date]['tax_free_pension']=values[9].value  #Total Deduction

        elif(str(values[8].value).replace(" ", "") == "Take home".replace(" ", "")):
             data[str(values[4].value)+date]['take_home']=values[9].value  #less_takehome

        # elif(str(values[8].value).replace(" ", "") == "NULL".replace(" ", "")):

        elif(str(values[8].value).replace(" ", "") == "Net pay".replace(" ", "")):
             data[str(values[4].value)+date]['take_home']=values[9].value  #less_takehome

        elif(str(values[8].value).replace(" ", "") == "Total Income".replace(" ", "")):
            data[str(values[4].value)+date]['salary']=values[9].value  #salary

        elif(str(values[8].value).replace(" ", "") == "Net Tax".replace(" ", "")):
            data[str(values[4].value)+date]['net_tax']=values[9].value  #Net Tax

        elif(str(values[8].value).replace(" ", "") == "PAYE".replace(" ", "")):
            data[str(values[4].value)+date]['paye']=values[9].value  #salary

        elif(str(values[8].value).replace(" ", "") == "Basic Pay".replace(" ", "")):
            data[str(values[4].value)+date]['salary']=values[9].value  #salary

        elif(str(values[8].value).replace(" ", "") == "Net Basic".replace(" ", "")):
            data[str(values[4].value)+date]['gross_salary']=values[9].value  #salary

        elif(str(values[8].value).replace(" ", "") == "NSSF".replace(" ", "")):
            data[str(values[4].value)+date]['salary']=values[9].value  #salary

        elif(str(values[8].value).replace(" ", "") == "Normal O/T Amount".replace(" ", "")):
            data[str(values[4].value)+date]['o/t']=values[9].value  #o/t


        elif(str(values[8].value).replace(" ", "") == "Arrears".replace(" ", "")):
            data[str(values[4].value)+date]['arrears']=values[9].value  #arrears

        elif(str(values[8].value).replace(" ", "") == "Leave Balance".replace(" ", "")):
            data[str(values[4].value)+date]['leave_balance']=values[9].value  #leave balance

        elif(str(values[8].value).replace(" ", "") == "Teller/Back Off".replace(" ", "")):
            data[str(values[4].value)+date]['teller_back_office']=values[9].value  #leave balance

        elif(str(values[8].value).replace(" ", "") == "S0263.0001.0032".replace(" ", "")):
            data[str(values[4].value)+date]['helsb']=values[9].value  #helsb

        elif(str(values[8].value).replace(" ", "") == "S2476.0032.0072".replace(" ", "")):
            data[str(values[4].value)+date]['helsb']=values[9].value  #helsb


        elif(str(values[8].value).replace(" ", "") == "E0015.3226.0122".replace(" ", "")):
            data[str(values[4].value)+date]['helsb']=values[9].value  #helsb


        elif(str(values[8].value).replace(" ", "") == "S0302.0218.0102".replace(" ", "")):
            data[str(values[4].value)+date]['helsb']=values[9].value  #helsb


        elif(str(values[8].value).replace(" ", "") == "S0208.0041.0052".replace(" ", "")):
            data[str(values[4].value)+date]['helsb']=values[9].value  #helsb


        elif(str(values[8].value).replace(" ", "") == "S0306.0193.0112".replace(" ", "")):
            data[str(values[4].value)+date]['helsb']=values[9].value  #helsb


        elif(str(values[8].value).replace(" ", "") == "S0222.0156.0102".replace(" ", "")):
            data[str(values[4].value)+date]['helsb']=values[9].value  #helsb


        elif(str(values[8].value).replace(" ", "") == "Task Allowance".replace(" ", "")):
            data[str(values[4].value)+date]['task_allowances']=values[9].value  #task_allowances

        elif(str(values[8].value).replace(" ", "") == "S1818.0246.0102".replace(" ", "")):
            data[str(values[4].value)+date]['helsb']=values[9].value  #helsb


        elif(str(values[8].value).replace(" ", "") == "S0756.0003.0092".replace(" ", "")):
            data[str(values[4].value)+date]['helsb']=values[9].value  #helsb


        elif(str(values[8].value).replace(" ", "") == "S0195.0094.0122".replace(" ", "")):
            data[str(values[4].value)+date]['helsb']=values[9].value  #helsb


        elif(str(values[8].value).replace(" ", "") == "S0731.0115.0082".replace(" ", "")):
            data[str(values[4].value)+date]['helsb']=values[9].value  #helsb


        elif(str(values[8].value).replace(" ", "") == "Normal Overtime Hrs".replace(" ", "")):
            data[str(values[4].value)+date]['normal_overtime']=values[9].value  #normal_overtime


        elif(str(values[8].value).replace(" ", "") == "Sundays O/T Hours".replace(" ", "")):
            data[str(values[4].value)+date]['sunday_overtime']=values[9].value  #sunday_overtime

        elif(str(values[8].value).replace(" ", "") == "WeekDay Overtime @1.5".replace(" ", "")):
            data[str(values[4].value)+date]['weekday_overtime']=values[9].value  #weekday_overtime


        elif(str(values[8].value).replace(" ", "") == "Sundays Overtime @.2".replace(" ", "")):
            data[str(values[4].value)+date]['sunday_overtime2']=values[9].value  #sunday_overtime2


        elif(str(values[8].value).replace(" ", "") == "Rounding".replace(" ", "")):
            data[str(values[4].value)+date]['rounding']=values[9].value  #rounding

        elif(str(values[8].value).replace(" ", "") == "S0264.0062.0082".replace(" ", "")):
            data[str(values[4].value)+date]['helsb']=values[9].value  #helsb


        elif(str(values[8].value).replace(" ", "") == "Acting Allowanc".replace(" ", "")):
            data[str(values[4].value)+date]['acting_allowance']=values[9].value  #acting_allowance

        elif(str(values[8].value).replace(" ", "") == "S1626.0095.0112".replace(" ", "")):
            data[str(values[4].value)+date]['helsb']=values[9].value  #helsb


        elif(str(values[8].value).replace(" ", "") == "Disturbance All".replace(" ", "")):
            data[str(values[4].value)+date]['disturbance']=values[9].value  #helsb


        elif(str(values[8].value).replace(" ", "") == "N/Shift Allowan".replace(" ", "")):
            data[str(values[4].value)+date]['shift_allowance']=values[9].value  #helsb

        elif(str(values[8].value).replace(" ", "") == "Transport Allow".replace(" ", "")):
             data[str(values[4].value)+date]['transport_allow']=values[9].value  #transport_allow

        elif(str(values[8].value).replace(" ", "") == "S0423.0180.0062".replace(" ", "")):
            data[str(values[4].value)+date]['helsb']=values[9].value  #helsb

        elif(str(values[8].value).replace(" ", "") == "Unpaid Leave".replace(" ", "")):
            data[str(values[4].value)+date]['unpdaid_leave']=values[9].value  #unpdaid_leave

        elif(str(values[8].value).replace(" ", "") == "O/Stand Leave".replace(" ", "")):
            data[str(values[4].value)+date]['o/stand_leave']=values[9].value  #o/stand_leave

        elif(str(values[8].value).replace(" ", "") == "Leave Allowance".replace(" ", "")):
            data[str(values[4].value)+date]['leave_allowances']=values[9].value  #allowances

        elif(str(values[8].value).replace(" ", "") == "S0960.0313.0092".replace(" ", "")):
            data[str(values[4].value)+date]['helsb']=values[9].value  #helsb

        elif(str(values[8].value).replace(" ", "") == "House Rent".replace(" ", "")):
            data[str(values[4].value)+date]['house_rent']=values[9].value  #house rent

        elif(str(values[8].value).replace(" ", "") == "Air Ticket Allo".replace(" ", "")):
            data[str(values[4].value)+date]['air_ticket']=values[9].value  #air_ticket

        elif(str(values[8].value).replace(" ", "") == "Hse Loan/Rent".replace(" ", "")):
            data[str(values[4].value)+date]['hse_loan']=values[9].value  #hse_loan

        elif(str(values[8].value).replace(" ", "") == "Vehicle Benefit".replace(" ", "")):
            data[str(values[4].value)+date]['vehicle_benefit']=values[9].value  #vehicle_benefit

        elif(str(values[8].value).replace(" ", "") == "Bonus".replace(" ", "")):
            data[str(values[4].value)+date]['bonus']=values[9].value  #bonus

        elif(str(values[8].value).replace(" ", "") == "ADVANCE".replace(" ", "")):
            data[str(values[4].value)+date]['advance']=values[9].value  #advance

        elif(str(values[8].value).replace(" ", "") == "Mid Month Advances".replace(" ", "")):
            data[str(values[4].value)+date]['mid_month_advance']=values[9].value  #mid_month_advance

        elif(str(values[8].value).replace(" ", "") == "Sundays O/T Amount".replace(" ", "")):
            data[str(values[4].value)+date]['sunday_overtime']=values[9].value  #sunday_overtime

        elif(str(values[8].value).replace(" ", "") == "Services Award".replace(" ", "")):
            data[str(values[4].value)+date]['service_award']=values[9].value  #service_award

        elif(str(values[8].value).replace(" ", "") == "Share All. Pymt".replace(" ", "")):
            data[str(values[4].value)+date]['share_all_payment']=values[9].value  #share_all_payment

        elif(str(values[8].value).replace(" ", "") == "SalaryPaid".replace(" ", "")):
            data[str(values[4].value)+date]['salary']=values[9].value  #salary

        # elif(str(values[8].value).replace(" ", "") == "E".replace(" ", "")):

        elif(str(values[8].value).replace(" ", "") == " Add:  Loan benefit".replace(" ", "")):
            data[str(values[4].value)+date]['loan_benefit']=values[9].value  #loan_benefit

        elif(str(values[8].value).replace(" ", "") == "Leavel&O/Stand".replace(" ", "")):
            data[str(values[4].value)+date]['level_i/o']=values[9].value  #loan_benefit

        elif(str(values[8].value).replace(" ", "") == "Long Service".replace(" ", "")):
            data[str(values[4].value)+date]['long_service']=values[9].value  #loan_benefit

        elif(str(values[8].value).replace(" ", "") == "Recognition".replace(" ", "")):
            data[str(values[4].value)+date]['recognition']=values[9].value  #loan_benefit

        elif(str(values[8].value).replace(" ", "") == "TRA Cost".replace(" ", "")):
            data[str(values[4].value)+date]['tra_cost']=values[9].value  #loan_benefit

        elif(str(values[8].value).replace(" ", "") == "Unpaid Days".replace(" ", "")):
            data[str(values[4].value)+date]['unpaid_days']=values[9].value  #loan_benefit

        # data[str(values[4].value)+date]['pension_employee']=values[2].value  #pension_employee
        # data[str(values[4].value)+date]['pension_employer']=values[2].value  #pension_employer
        # data[str(values[4].value)+date]['medical_employee']=values[2].value  #medical_employee
        # data[str(values[4].value)+date]['medical_employer']=values[2].value  #medical_employer
        # data[str(values[4].value)+date]['taxdue']=values[2].value  #taxdue
        # data[str(values[4].value)+date]['meals']=values[2].value  #meals
        # data[str(values[4].value)+date]['department']=values[2].value  #department
        # data[str(values[4].value)+date]['position']=values[2].value  #position
        # data[str(values[4].value)+date]['branch']=values[2].value  #branch
        # data[str(values[4].value)+date]['pension_fund']=values[2].value  #pension_fund
        # data[str(values[4].value)+date]['membership_no']=values[2].value  #membership_no
        # data[str(values[4].value)+date]['sdl']=values[2].value  #sdl
        # data[str(values[4].value)+date]['wcf']=values[2].value  #wcf
        # data[str(values[4].value)+date]['less_takehome']=values[2].value  #less_takehome
        data[str(values[4].value)+date]['date']=date  #due_date
        # data[str(values[4].value)+date]['payroll_date']=date  #payroll_date
    #     data[str(values[4].value)+date]['bank']=values[2].value  #bank
    #     data[str(values[4].value)+date]['bank_branch']=values[2].value  #bank_branch
    #     data[str(values[4].value)+date]['account_no']=values[2].value  #account_no
    #     data[str(values[4].value)+date]['created_at']=values[2].value  #created_at
    k+=1
print(k)

file = "test.xlsx" #load the work book
wb_obj = load_workbook(filename = 'pension payments.xlsx')
wsheet = wb_obj['Sheet1']
dataDict = {}

# data={}

#ID	CompanyCode	FISCALYEAR	Mon	MonthName	CODENO	Group	GroupNo	Position	Desc	Value	Name	Department	Description	BANKNAME	BRANCH	PayMethod	NPFNUMBER	JobTitle	msg1	msg2	msg3	Wstation	WStation_Name	Hired_Date	STATUS	upsize_ts
#0	 1	              2   	3	       4	5	      6	       7    	8	     9	        10	11	      12	     13	           14	      15	16	           17	       18	     19	      20	21	      22	      23	            24	       25	26



#id	empID	salary	allowances	pension_employee	pension_employer	medical_employee	medical_employer	taxdue	meals	department	position	branch	pension_fund	membership_no	sdl	wcf	less_takehome	due_date	payroll_date	bank	bank_branch	account_no	created_at	updated_at
#1   	2	 3	     4	         5	                  6	                  7	                    8	              9	     10	     11	         12	          13	    14	              15         16	 17 	18	        19	        20	            21	        22	                    23	        24	25
k=0
r=0
temp=[]
while r<30:
    temp.append('')
    r+=1
    # print(r)
for key, *values in wsheet.iter_rows():
    temp={}
    if(k>=1):
        # print(values[2].value)

#ID	CompanyCode	CODENO	FISCALYEAR	MON	Month	LASTDAY	NAME	PAYDOLLAR	EXCHANGE	NSSF Number	Income	Employee Cotribution	Employer Contribution	Total	Receipt Number	Receipt Date	NPFPPS	upsize_ts
#       0          1	       2     3 	   4	     3	  6	       7	    8	         9	            10	              11                        12	      13	    14	         15  	        16	      17	                      9

        # # date='25/'+str(values[2].value)+'/'+str(values[1].value)
        date='25-'+str(values[3].value)+'-'+str(values[2].value)
        # print(date)
        # # for value in values:
        if str(values[1].value)+date in data:

            # data[str(values[1].value)+date]=temp

            # data[str(values[1].value)+date]['date']=date
            data[str(values[1].value)+date]['employer_contribution']=values[11].value
            data[str(values[1].value)+date]['employee_contribution']=values[12].value
            data[str(values[1].value)+date]['membership_no']=values[9].value
        # else:
        #     print(date)
    k=k+1
print(k)

i=0
# for key in data:
    # print(data[key])
    # print(k)
# for key in data:
#     print(data[key])
    # if(str(values[8].value).replace(" ", "") == "i+=1.replace(" ", "")"):

# prelse int(i)
# {'taxable_gross': '0', 'gross_pay': '0', 'total_deduction': 740700, 'tax_free_pension': 230000, 'take_home': 1559300, 'salary': '0', 'net_tax': '0', 'paye': '0',
#  'gross_salary': '0', 'o/t': '0', 'arrears': '0', 'leave_balance': '0', 'teller_back_office': '0', 'helsb': '0', 'task_allowances': '0', 'normal_overtime': '0', 
#  'sunday_overtime': '0', 'weekday_overtime': '0', 'sunday_overtime2': '0', 'rounding': '0', 'acting_allowance': '0', 'disturbance': '0', 'shift_allowance': '0',
#   'transport_allow': '0', 'unpdaid_leave': '0', 'o/stand_leave': '0', 'leave_allowances': '0', 'house_rent': '0', 'air_ticket': '0', 'hse_loan': '0', 
#   'vehicle_benefit': '0', 'bonus': '0', 'advance': '0', 'mid_month_advance': '0', 'service_award': '0', 'share_all_payment': '0', 'loan_benefit': '0', 
#   'level_i/o': '0', 'long_service': '0', 'recognition': '0', 'tra_cost': '0',
#  'unpaid_days': '0', 'employer_contribution': 230000, 'employee_contribution': 230000, 'membership_no': 50045301, 'date': '25/11/2013', 'empID': 100223}

#id	empID	salary	allowances	pension_employee	pension_employer	medical_employee	medical_employer	taxdue	meals	department	position	branch	pension_fund	membership_no	sdl	wcf	less_takehome	due_date	payroll_date	bank	bank_branch	account_no	created_at	updated_at
#1   	2	 3	     4	         5	                  6	                  7	                    8	              9	     10	     11	         12	          13	    14	              15         16	 17 	18	        19	        20	            21	        22	                    23	        24	25

import csv
i=0
with open("test.csv", "wt") as fp:
    writer = csv.writer(fp, delimiter=",")
    # writer.writerow(["your", "header", "foo"])  # write header
    for key in data:
        print(data[key])
        if(data[key]['membership_no'] is None):
            data[key]['membership_no']=0
        if(data[key]['membership_no'] == 'Null'):
            data[key]['membership_no']=0

        writer.writerow([
        i,      #id
        data[key]['empID'],      #empID
        data[key]['salary'],      #salary
        data[key]['shift_allowance'],      #allowances
        data[key]['employee_contribution'],      #pension_employee
        data[key]['employer_contribution'],      #pension_employer
        data[key]['shift_allowance'],      #medical_employee
        data[key]['shift_allowance'],      #medical_employer
        data[key]['shift_allowance'],      #taxdue
        data[key]['shift_allowance'],      #meals
        data[key]['shift_allowance'],      #department
        data[key]['shift_allowance'],      #position
        data[key]['shift_allowance'],      #branch
        data[key]['shift_allowance'],      #pension_fund
        data[key]['membership_no'],      #membership_no
        data[key]['shift_allowance'],      #sdl
        data[key]['shift_allowance'],      #wcf
        data[key]['take_home'],      #less_takehome
        data[key]['date'],      #due_date
        data[key]['date'],      #payroll_date
        data[key]['shift_allowance'],      #bank
        data[key]['shift_allowance'],      #bank_branch
        data[key]['shift_allowance'],      #account_no
        data[key]['date'],
        data[key]['date']]    #created_at
       
        )
        i=i+1